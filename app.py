import pandas as pd
import tkinter as tk
import openpyxl
import qpl_app, time, os
from tkinter import *
from tkinter import filedialog

class Application(Frame):

    def __init__(self,master):

        Frame.__init__(self,master)
        self.master = master

        self.qpl_file_path = tk.StringVar()
        self.app_tracker_file_path = tk.StringVar()
        self.checklist_file_path = tk.StringVar()
        self.sfchecklist_file_path = tk.StringVar()
        self.save_file_path = tk.StringVar()
        self.oemManufacturer = tk.StringVar()
        self.privateLabelManufacturer = tk.StringVar()
        self.manufacturer = tk.StringVar()
        self.manufacturers = []


        self.qplLabel = tk.Label(self.master,text='QPL Download File:',font=18).grid(row = 0, column = 0)
        self.qplPathLabel = tk.Label(self.master)
        self.qplButton = Button(text = 'Open File', command = self.openQPLFile).grid(row = 0, column = 3)

        self.appLabel = Label(self.master,text='Application Tracker:',font=18).grid(row = 1, column = 0)
        self.appPathLabel = Label(self.master)
        self.appButton = Button(text = 'Open File', command = self.openAppFile).grid(row = 1, column = 3)

        self.oemLabel = Label(self.master,text='OEM Manufacturer:',font=18).grid(row = 2, column = 2)
        self.privateLabel = Label(self.master,text = 'Private Label Manufacturer:',font = 18).grid(row = 6, column = 2)

        self.checkListLabel = Label(self.master,text = 'Private Label Check List:',font = 18).grid(row = 9, column = 0)
        self.checkListPathLabel = Label(self.master)
        self.checkListButton = Button(text = 'Open File', command = self.openCheckListFile).grid(row = 9, column = 3)
        self.startButton = Button(text = 'Start', command = self.importData).grid(row = 10, column = 1,columnspan = 2)

        self.sfCheckListLabel = Label(self.master,text = 'Single Family Check List:',font = 18).grid(row = 11, column = 0)
        self.sfCheckListPathLabel = Label(self.master)
        self.sfCheckListButton = Button(text = 'Open File', command = self.openSFCheckListFile).grid(row = 11, column = 3)
        self.sfStartButton = Button(text = 'Start', command = self.importDataSF).grid(row = 12, column = 1,columnspan = 2)

        # self.pack()
        self.create_widgets()

    def create_widgets(self):

        #qpl
        self.search_var_qpl = StringVar()
        self.search_var_qpl.trace("w", self.update_list)
        self.entry_qpl = Entry(self.master, textvariable = self.search_var_qpl, width = 45)
        self.lbox_qpl = Listbox(self.master, width = 45, height = 10,exportselection = 0)

        self.entry_qpl.grid(row = 4, column = 2, padx = 10, pady = 3)
        self.lbox_qpl.grid(row = 5, column = 2, padx = 10, pady = 3)

        #private label
        self.search_var_app = StringVar()
        self.search_var_app.trace("w", self.update_list)
        self.entry_app = Entry(self.master, textvariable = self.search_var_app, width = 45)
        self.lbox_app = Listbox(self.master, width = 45, height = 10,exportselection = 0)

        self.entry_app.grid(row = 7, column = 2, padx = 10, pady = 3)
        self.lbox_app.grid(row = 8, column = 2, padx = 10, pady = 3)

    def openQPLFile(self,*args):

        self.qpl_file_path = filedialog.askopenfilename(initialdir = 'P:\\DesignLights\\Admin\\Application Tools\\_Static QPL\\New Website QPL Download')
        self.qplPathLabel.config(text = os.path.split(self.qpl_file_path)[1])
        self.qplPathLabel.grid(row = 0, column = 1,columnspan = 2)
        self.manufacturers = qpl_app.getManufacturers(self.qpl_file_path)
        self.manufacturer = self.manufacturers[0]

        self.update_list()

    def openAppFile(self,*args):
        self.app_tracker_file_path = filedialog.askopenfilename(initialdir = 'P:\\DesignLights\\Admin\\Application Tools\\_Static QPL\\Combined Apps Download (New and Old Website)')
        self.appPathLabel.config(text = os.path.split(self.app_tracker_file_path)[1])
        self.appPathLabel.grid(row = 1, column = 1,columnspan = 2)


    def openCheckListFile(self,*args):

        self.checklist_file_path = filedialog.askopenfilename(initialdir = 'P:\\DesignLights\\Admin\\Application Tools\\Review Checklists\\Private_Label_Checklist')
        self.checkListPathLabel.config(text = os.path.split(self.checklist_file_path)[1])
        self.checkListPathLabel.grid(row = 9, column = 1,columnspan = 2)


    def openSFCheckListFile(self,*args):
        self.sfchecklist_file_path = filedialog.askopenfilename(initialdir = 'P:\\DesignLights\\Admin\\Application Tools\\Review Checklists\\Combined_Single-Family_Checklist')
        self.sfCheckListPathLabel.config(text = os.path.split(self.sfchecklist_file_path)[1])
        self.sfCheckListPathLabel.grid(row = 11, column = 1,columnspan = 2)


    def update_list(self, *args):
        search_term_qpl = self.search_var_qpl.get()

        # Just a generic list to populate the listbox
        lbox_list_qpl = self.manufacturers

        self.lbox_qpl.delete(0, END)

        for item in lbox_list_qpl:
            if search_term_qpl.lower() in item.lower():
                self.lbox_qpl.insert(END, item)

        search_term_app = self.search_var_app.get()

        # Just a generic list to populate the listbox
        lbox_list_app = self.manufacturers

        self.lbox_app.delete(0, END)

        for item in lbox_list_app:
            if search_term_app.lower() in item.lower():
                self.lbox_app.insert(END, item)

    def importData(self,*arg):

        self.oemManufacturer = self.lbox_qpl.get(self.lbox_qpl.curselection())
        self.privateLabelManufacturer = self.lbox_app.get(self.lbox_app.curselection())

        df= qpl_app.oemProducts(self.qpl_file_path,self.oemManufacturer)

        with open(self.checklist_file_path,'a'):
            book = openpyxl.load_workbook(self.checklist_file_path)
            writer = pd.ExcelWriter(self.checklist_file_path,engine = 'openpyxl')
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            df.to_excel(writer, 'OEM Products', index = False)
            writer.save()

        df = qpl_app.plProducts(self.qpl_file_path,self.privateLabelManufacturer)

        with open(self.checklist_file_path,'a'):
            book = openpyxl.load_workbook(self.checklist_file_path)
            writer = pd.ExcelWriter(self.checklist_file_path,engine = 'openpyxl')
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            df.to_excel(writer, 'PL Products', index = False)
            writer.save()

        df = qpl_app.app_tracker(self.app_tracker_file_path,self.oemManufacturer)

        with open(self.checklist_file_path,'a'):
            book = openpyxl.load_workbook(self.checklist_file_path)
            writer = pd.ExcelWriter(self.checklist_file_path,engine = 'openpyxl')
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            df.to_excel(writer, 'OEM Applications', index = False)
            writer.save()


    def importDataSF(self,*args):
        self.oemManufacturer = self.lbox_qpl.get(self.lbox_qpl.curselection())
        df = qpl_app.oemProducts(self.qpl_file_path,self.oemManufacturer)

        with open(self.sfchecklist_file_path,'a'):
            book = openpyxl.load_workbook(self.sfchecklist_file_path)
            writer = pd.ExcelWriter(self.sfchecklist_file_path,engine = 'openpyxl')
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            df.to_excel(writer, 'OEM Products', index = False)
            writer.save()


root = tk.Tk()
root.geometry('600x600')
root.title('DLC Application Tracker and QPL Checklist')
app = Application(master = root)
app.mainloop()
